# git add "folder/filename"
# git commit -m "Commit message"
# git push -u origin main

import time
import openpyxl
import pandas as pd
import random
import os
import itertools
import threading
import sys
import pyinputplus as pyplus

# Name all global variables
global difficulty
global profile
global name
global enemy_profile
global player_HP
global floor
global enemycomment 
global abilities
global player_lvl
global player_gold
global critrate
global ID
global upgrade
global cont_choose_prof
global done

# Ivan
def userattackinput(timer):
    # Allows user to enter multiple attack words within a time limit
    userinputstring("\nPress ENTER key to continue")
    print(f"\n{profile[1]} you have to type as many attack words as you can within {timer} seconds!!")
    time.sleep(.5)
    print("3...")
    time.sleep(1)
    print("2...")
    time.sleep(1)
    print("1...")
    time.sleep(1)
    print("TYPE\n")
    start_time = time.time()
    last_time = start_time
    total_time = 0
    words = []

    # While got time, user can enter as many words as they want
    while total_time < timer:
        word = userinputstring("")
        if word != "":
            # To improve game experience, DO NOT COPY PASTE
            words.append(word)
        total_time = round((time.time() - last_time), 2)
    print("\nWords typed: "+str(len(words)))
    return words

# Ivan
def userinputstring(comment):
    # loop continues until a valid string is input
    cont = True
    while cont:
        try:
            # take user input as a string
            value = str(input(comment)).strip()
            # end the loop if the input is valid
            cont = False
            return value
        except ValueError:
            # if the input is not a valid string, the loop continues
            continue

# Ivan
def animate():
    # Animate loading animation
    for c in itertools.cycle(['⡀', '⡄', '⡆', '⡇', '⣇', '⣧', '⣷', '⣿']):
        if done:
            break

        # Type out "Loading" while another tasks is working in another thread
        sys.stdout.write('\rLoading ' + c)
        sys.stdout.flush()
        time.sleep(0.1)

# Ivan
def userinputint(comment):
    # Validate integer input with a loop
    cont = True
    while cont:
        try:
            # Get input and convert to integer
            value = int(input(comment))
            cont = False
        except ValueError:
            # Handle ValueError exception and continue the loop
            continue
    return value

# Xuan Yu
def getprofile(id, sheet):
    # Empty list for profile
    global profile

    profile = []
    
    # Loop through rows in sheet
    exist = False
    for i, row in enumerate(sheet):

        # If the value of index is 0 "Column name", skip the row
        if i == 0:
            continue

        # If the ID is equal to the ID required
        if row[0].value == id:
            exist = True

            # Loop through columns and append values to profile list
            for cols in sheet.iter_cols():
                x = cols[i].value
                profile.append(x)
   
    # Return if profile exists or not
    if exist:
        update()
        return True
    if not exist:
        return False

# Xuan Yu
def update():

    # Update profile list
    global difficulty
    global player_HP
    global player_lvl
    global floor
    global name
    global player_gold
    global critrate
    global ID
    global upgrade
    global wordstyped
    global validwordstyped
    global losses
    global wins
    global timetaken

    # record the player and update in excel list
    ID = profile[0]
    name = profile[1]
    difficulty = profile[2] 
    player_HP = profile[3]
    player_lvl = profile[4]
    player_gold = profile[5]
    floor = profile[6]
    critrate = profile[7]
    upgrade = profile[8]
    wordstyped = profile[9]
    validwordstyped = profile[10]
    losses = profile[11]
    wins = profile[12]
    timetaken = profile[13]

# Xuan Yu
def checkprofile():

    # Print out a list of player's profile
    print(f"\n{profile[1]}'s profile:\n--------------------------------\n")
    time.sleep(.5)
    print(f"Current difficulty   >>>   {difficulty}") # Difficulty
    time.sleep(.5)
    print(f"Current HP   >>>   {player_HP}") # HP
    time.sleep(.5)
    print(f"Current Level   >>>   {player_lvl}") # Level
    time.sleep(.5)
    print(f"Current Gold   >>>   {player_gold}") # Gold
    time.sleep(.5)
    print(f"Current Crit rate   >>>   {critrate*100}%")  # Crit Rate
    time.sleep(.5)
    if floor<=5:
        print(f"Current floor   >>>   {floor}") # Floor
        time.sleep(.5)
    print(f"Words typed   >>>   {wordstyped}")  # words typed
    time.sleep(.5)
    print(f"Valid words typed   >>>   {validwordstyped}")  # valid words typed
    time.sleep(.5)
    print("Accuracy   >>> {0:10.2%}".format(validwordstyped/wordstyped))  # accuracy
    time.sleep(.5)
    print(f"Wins   >>>   {wins}")  # wins
    time.sleep(.5)
    print(f"Losses   >>>   {losses}")  # losses
    time.sleep(.5)
    print("Average speed   >>> {0:10.2f} words/min".format(validwordstyped/timetaken*60))  # speed
    time.sleep(3)
    print(f"\n\nAbilities unlocked")
    time.sleep(.5)
    printabilities("collection")
    upgrade = getartifacts("collection")
    time.sleep(1)
    # print out the reward when player upgrade
    if len(upgrade) != 0:
        print(f"\nThroughout your journey, you have obtained {len(upgrade)} artifacts:")
        i = 1
        #print out the upgrade details and record 
        for x in upgrade:
            print(f"{i}. {x}")
            time.sleep(.2)
            i += 1
     # when player no upgrade       
    else:
        print("\nArtifacts are awaiting for you to be uncovered...")

    userinputstring("\n\nPress ENTER to continue")

# Yuven
def getenemyprofile():
    global enemy_profile
    global floor
    enemy_profile = []
    global enemyatck
    enemyatck = {}
    enemyframe = openpyxl.load_workbook("Enemies.xlsx")
    enemysheet = enemyframe.worksheets[floor]
    listingenemyprofile = lambda a :  enemy_profile.append(enemysheet[a].value)
    listingenemyprofile("A1") # Name
    listingenemyprofile("B1") # HP
    listingenemyprofile("B1") # Max HP
    listingenemyprofile("K1") # Timer
    listingenemyprofile("C1") # Gold

    # alternative

    #enemy_profile.append(enemysheet["A1"].value) # enemy name
    #enemy_profile.append(enemysheet["B1"].value) # enemy HP
    #enemy_profile.append(enemysheet["B1"].value) # enemy max HP
    #enemy_profile.append(enemysheet["K1"].value) # enemy timer
    #enemy_profile.append(enemysheet["C1"].value) # enemy gold

    enemybackground = enemysheet["E1"].value # get enemy background
    for row in enemysheet:
        enemyatck[row[5].value] = row[7].value, row[6].value
    print("\n")
    print(enemybackground)
    time.sleep(2)
    print("\n")

    # Get enemy comment
    global enemycomment
    enemycomment = []
    for row in enemysheet:
        enemycomment.append(row[9].value)

    for num, row in enumerate(enemysheet):
        print(f"{num+1}. {row[3].value}")
        time.sleep(0.2)
    #print(enemyatck)

# Ivan
def combat(gamefloor):
    global player_HP
    global cont_choose_prof
    global done
    player_maxHP = player_HP # set maximum HP of player before combat
    gameover = False
    playerwin = False
    firsttime = False

    """
    Check if first time playing the floor
    If found to be 1st time, they will only get golds after the round
    """

    if floor == profile[6]:
        firsttime = True
    else:
        firsttime = False

    if gamefloor == 0:
        print("\n\nTutorial session starting...")
    
    while not gameover:
        # check if both player or enemy alive
        player_game_HP = player_maxHP # Reset HP
        winner = ""
        playeralive = True
        enemyalive = True
        getenemyprofile()
        timer = enemy_profile[3]
        profile[13] += timer # timetaken
        time.sleep(2)
        print("")
        userinputstring("Press ENTER key to continue\n")
        rounds = 1

        # Loop the combat while both are alive
        while playeralive and enemyalive:
            os.system("cls")
            print(f"--------\nRound {rounds}\n--------")

            """
            This part is for the PLAYER
            """

            printabilities("combat") # get list of valid abilities
            wordlist = userattackinput(timer) # Get user input
            profile[9] += len(wordlist) # Words typed
            validlist = []

            for x in abilities:
                for y in wordlist:

                    # Set difficulty
                    if difficulty == "Easy":
                       newx = x.lower().replace(" ", "") # remove space and case
                       newy = y.lower().replace(" ", "") # remove space and case
                    if difficulty == "Hard": # Remain the same case and space
                        newx = x
                        newy = y
                    if newx == newy :
                        # append attack words in list
                        validlist.append(x)
            profile[10] += len(validlist) # Valid words typed
            playertotaldmg = 0
            playertotalheal = 0
            healaccumulated = 0
            i = 1

            """
            check valid attack words
            for example:
                2 fireball dmg 100
                1 stupefy regen 200
            
            """

            time.sleep(1)
            print(f"\nValid words typed: {len(validlist)}")
            # Print out all the valid words typed by user
            for x in validlist:
                print(f"{i}. {x} -- {abilities[x][0]} {abilities[x][1]}")
                time.sleep(0.2)
                # add all dmg and regen
                i += 1

                # Check the type: dmg//regen
                if abilities[x][0] == "dmg":
                    playertotaldmg += abilities[x][1]
                if abilities[x][0] == "regen":
                    playertotalheal += abilities[x][1]
            print(f"\nTotal damage accumulated: {playertotaldmg}")

            """
            RULE 1: Check if HP is smaller than 80%
            If > 80% can't heal
            display error message
            """

            if playertotalheal > 0:
                if player_game_HP < (player_maxHP*0.8):
                    player_newHP = player_game_HP + playertotalheal
                    healaccumulated = playertotalheal
                    
                    """
                    RULE 2: Make sure added HP doesn't exceed the max HP
                    If exceeds, return the max Hp of player
                    """

                    if player_newHP > player_maxHP:
                        healaccumulated = player_newHP - player_game_HP
                        player_game_HP = player_maxHP
                    else:
                        player_game_HP = player_newHP
                    print(f"Total heal accumulated: {healaccumulated}")
                else:
                    print(f"You can't heal, your HP is more than 80%")
            time.sleep(1)
                
            # Check if crit damage applied
            rand = random.random() # generate random from 0-1
            if critrate > rand:
                playertotaldmg = round(playertotaldmg*1.5)
                print(f"Critical attack!!!\nTotal damage: {playertotaldmg}")

            # Deliver dmg to enemy
            if playertotaldmg > 0:
                enemy_profile[1] -= playertotaldmg
                if enemy_profile[1] <= 0: # Check if enemy is dead
                    print(f"\n{enemy_profile[0]} has been defeated!")
                    enemyalive = False
                    winner = profile[1] # Define the winner
                    gameover = True # Break game loop
                    playerwin = True
                    break

            if playertotaldmg<=0 and healaccumulated<= 0:
                print(f"\n{name} did not use any attacks or heal this turn.")

            # print status bar
            print(f"\n{name}'s HP {player_game_HP}/{player_maxHP}  ||  {enemy_profile[0]}'s HP {enemy_profile[1]}/{enemy_profile[2]}\n")

            """
            This part is for the ENEMY
            """

            time.sleep(1)
            enemymove = getranatk()  # get random attack from enemy
            time.sleep(2)
            print(f"\nIt is {enemy_profile[0]}'s turn\n")
            time.sleep(1)
            print(random.choice(enemycomment)) # Randomly prints out a message from the enemy
            time.sleep(1)
            print(f"{enemy_profile[0]} used {enemymove[0]} and {enemymove[2]} {enemymove[1]} points") # status bar

            """
            Check if enemy move is attack or regen
            If attack: playerhp - enemy attack
            If regen: enemyhp + enemy regen
            """

            if enemymove[2] == "attacked":
                player_game_HP -= enemymove[1] 
                if player_game_HP <= 0: # Check if player dead
                    time.sleep(1)
                    print("\nYou have died... Game Over")
                    playeralive = False
                    winner = enemy_profile[0]
                    gameover = True
                    break
            else:
                enemy_profile[1] += enemymove[1]
            time.sleep(1)

            # print status bar
            print(f"\n{profile[1]}'s HP {player_game_HP}/{player_maxHP}  ||  {enemy_profile[0]}'s HP {enemy_profile[1]}/{enemy_profile[2]}\n")
            time.sleep(2)
            userinputstring("Press ENTER key to continue\n")
        
            # repeat
            rounds += 1
        time.sleep(1)
        os.system("cls")
        print(f"\n{winner} Won!!!")

        # Ask for rematch?
        while playeralive == False:
            profile[11] += 1 # Losses
            contplay = pyplus.inputYesNo("Do you want a rematch? (Yes//No)\n>>> ",blank=False)
            if contplay == "yes":
                gameover = False
                continue
            elif contplay == "no":
                profilesheet = profileframe.worksheets[0]
                return chooseprofile(profilesheet)

        # Update profile if player wins
        if playerwin and firsttime and gamefloor<5:
            time.sleep(2)
            print(f"\n***Level Up!!***\n------------------\nLevel {player_lvl} --> Level {player_lvl+1}\n")
            profile[12] += 1 # Wins
            profile[4] += 1 # Level
            profile[5] += enemy_profile[4] # Gold
            profile[6] += 1 # Progress
            time.sleep(1)
            print(f"You have gained {enemy_profile[4]} golds from the battle\nTotal gold in piggy bank is {profile[5]}")
            upgrade = getartifacts("combat") # get upgrades from excel
            update()
            time.sleep(1)

            # check if the upgrade exists 
            if upgrade is not None:
                print(f"\nCongratulations you have unlocked {upgrade}\n")
                time.sleep(1)
            print(f"HP: {player_maxHP} --> {profile[3]}\n")
            time.sleep(1)
            print("New set of abilities unlocked!!\n")
            printabilities("combat") # print new abilities to be used on next floor

        # if gamefloor is last floor
        elif gamefloor == 5:
            time.sleep(2)
            print(f"\n***Level Up!!***\n------------------\nLevel {player_lvl} --> Level {player_lvl+1}\n")
            profile[4] += 1 # Level
            profile[5] += enemy_profile[4] # Gold
            profile[6] += 1 # Progress
            update()
            print(f"You have gained {enemy_profile[4]} golds from the battle\nTotal gold in piggy bank is {profile[5]}\n")
            
        # if playerwin only, applies to replay  
        elif playerwin:
            time.sleep(2)
            profile[12] += 1 # Wins
            profile[5] += enemy_profile[4] # Gold
            update()
            print(f"You have gained {enemy_profile[4]} golds from the battle\nTotal gold in piggy bank is {profile[5]}\n")
            
        print("\n\n")    

        # Start loading animation
        done = False
        t = threading.Thread(target=animate)
        t.start()

        # Update data in excel
        profilesheet = profileframe.worksheets[0]
        for num, row in enumerate(profilesheet):
            if num == 0:
                continue
            elif int(row[0].value) == profile[0]:
                row[3].value = profile[3] # max HP
                row[4].value = profile[4] # Lvl
                row[5].value = profile[5] # gold
                row[6].value = profile[6] # progress
                row[9].value = profile[9] # Words typed
                row[10].value = profile[10] # Valid words typed
                row[11].value = profile[11] # Defeats
                row[12].value = profile[12] # Wins
                row[13].value = profile[13] # Time taken
        profileframe.save("playerprofile.xlsx")
        done = True
        
        """
        Menu after combat, contains 2 conditions:
        1. gamefloor <= 5 (not the last)
        2. gamefloor = 6 (completed the last floor)
        """

        # Check if the previous floor is the last floor
        if gamefloor<=5:

            # Ask player if want to continue playing the game?
            contchoosecontgame = True
            while contchoosecontgame:
                os.system("cls")
                continuegame =  pyplus.inputMenu(["Proceed", "Buy Upgrades", "Back to Profile", "Quit Game"], "What do you want to do next?\n",blank=False)
                if continuegame == "Proceed":
                    contchoosecontgame = False
                    return True

                elif continuegame == "Quit Game":
                    contchoosecontgame = False
                    return False
            
                elif continuegame == "Buy Upgrades":
                    contchoosecontgame = True
                    buyupgrades()

                elif continuegame == "Back to Profile":
                    contchoosecontgame = False
                    return chooseprofile(profilesheet)
            

        else:
            os.system("cls")
            print("\nCongratulations you have completed the game")
            continuegame =  pyplus.inputMenu(["Back to Profile", "Quit Game"], "What do you want to do next?\n",blank=False)
            if continuegame == "Back to Profile":
                return chooseprofile(profilesheet)
            elif continuegame == "Quit Game":
                return False

# Ivan
def getartifacts(use):
    # Get upgrade list
    global player_lvl
    upgradesheet = profileframe.worksheets[2]
    player_lvl = profile[4]
    upgradelist = []

    """
    get the list of upgrades from excel
    2 types:
    1. combat --> returns only 1 
    2. collection --> returns all upgrades
    """

    for num, row in enumerate(upgradesheet):
        if use == "combat":
            if num == 0:
                continue

            # If playerlvl = required, HP increase
            elif int(row[1].value) == player_lvl:
                profile[3] += int(row[2].value)
                return row[0].value
            # If playerlvl < required, ignore
            elif int(row[1].value) > player_lvl:
                return None
        elif use == "collection":
            if num == 0:
                continue
            # Return all if playerlvl >= required
            elif player_lvl >= int(row[1].value):
                upgradelist.append(row[0].value)
    return upgradelist

# Yuven
def getranatk():
    # RULES:
    #   1. Healing can only be done when HP is smaller than 30%
    #   2. Any attack must be random
    global enemy_profile
    global enemyatck
    enemy_HP = enemy_profile[1]
    enemy_maxHP = enemy_profile[2]
    enemyattackwords = [x for x in enemyatck]
    contgetranatck = True
    while contgetranatck:
        # Rule 2
        randomattack = random.choice(enemyattackwords)
        # Calc dmg or regen
        if enemyatck[randomattack][0] == "dmg":
            enemydmgorhealvalue = enemyatck[randomattack][1]
            enemydmgorheal = "attacked"
            contgetranatck = False
        if enemyatck[randomattack][0] == "regen":
            # Rule 1
            if enemy_HP > enemy_maxHP*0.3:
                continue
            else:
                enemydmgorhealvalue = enemyatck[randomattack][1]
                enemydmgorheal = "healed"
                contgetranatck = False
    enemymove = [randomattack, enemydmgorhealvalue, enemydmgorheal]
    return enemymove

# Ivan
def print_data_in_table_form(file_path, sheet_name, range):
    df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=None, na_values=['NA'], usecols=range)
    print(df)

# Ivan
def choooseloadornew():
    global profile
    global name
    global player_HP
    global floor
    global difficulty
    global profileframe
    global cont_choose_prof
    global done

    # Start loading animation
    done = False
    t = threading.Thread(target=animate)
    t.start()

    # Load the profile from excel
    profileframe = openpyxl.load_workbook("playerprofile.xlsx")
    profilesheet = profileframe.worksheets[0]
    done = True

    cont_choose_loadornew = True # continue to choose load game // new game
    while cont_choose_loadornew:

        # creates a menu with a few choices
        if profilesheet.max_row == 1:
            choice = pyplus.inputMenu(["New Game", "Quit Game"],"\n",blank=False)
        
        else:
            choice = pyplus.inputMenu(["Load Game", "New Game", "Quit Game"],"\n",blank=False)

        if choice == "Load Game":
            cont_choose_prof = True 
            # continue to choose profiles
            return chooseprofile(profilesheet)

        elif choice == "New Game":
            if createprofile(profilesheet) == False:
                # Create a new profile, after that returns back to load/new game menu
                cont_choose_loadornew = True

            else:
                # assign values to new row of data
                rows = profilesheet.max_row
                done = False
                t = threading.Thread(target=animate)
                t.start()

                for i in range(14):
                    if i == 8:
                        continue
                    else:
                        profilesheet.cell(row=rows+1, column=i+1).value = profile[i]
                profileframe.save("playerprofile.xlsx")
                done = True
                os.system("cls")

        elif choice == "Quit Game":
            # Quit the game
            return False

# Ivan
def chooseprofile(profilesheet):
    global done
    global floor
    count = 0
    cont_choose_prof = True

    # While the game doesn't start, loop choose profile menu
    while cont_choose_prof:
        os.system("cls")
        print("\n-------------------------------------------------------\n\nThese profiles are available\n\n")
        time.sleep(.5)
        #print profile choices
        print_data_in_table_form("playerprofile.xlsx", 0, "A:G")

        print("\n0. Edit profile")
        time.sleep(.5)
        getid = True
        while getid:
            delete = False
            playerID = userinputstring('\nPlease choose a player ID or type "edit" before ID to edit the ID: ').split(" ")

            """
            Checks the input of user, format must be:
                1. command ID
                2. ID
            """

            if len(playerID) == 1:
                # Check if ID is valid
                try:
                    ID = int(playerID[0])
                    getid = False
                # Command and ID without separator
                except ValueError:
                    print("Remember to leave a space in between and include an ID")
                    continue

            if len(playerID) == 2:
                # Check if first word is command
                if playerID[0] == "edit":
                    # Check if ID exist
                    try:
                        ID = int(playerID[1])
                        exist = getprofile(ID, profilesheet)
                        if not exist:
                            print("ID doesn't exits")
                            continue
                        else:
                            getid = False
                    except ValueError:
                        print("Must include a valid ID")
                        continue
                # Command not recognised
                else:
                    print(f"Command '{playerID[0]}' is not recognised!!!")
                    continue

                # Check if profile is deleted
                delete = editprofile(ID, profilesheet)
                print("\n\n")
                done = False
                t = threading.Thread(target=animate)
                t.start()
                profileframe.save("playerprofile.xlsx")
                done = True
            continue

        # get profile using function
        if delete:
            os.system("cls")
            choooseloadornew()

        # Check if profile doesn't exists
        elif not getprofile(ID, profilesheet):
            count += 1

            if count == 3: #Machine fed up
                print(".")
                time.sleep(.5)
                print("..")
                time.sleep(.5)
                print("...")
                time.sleep(.5)
                print("....")
                time.sleep(2)
                print("Are you dumb or what??? I told you there's no such ID!!!")
                time.sleep(1)
                print("I DARE YOU TO DO THAT ONE MORE TIME...\n")
                time.sleep(1)
            if count == 4:
                print("Goodbye! We don't tolerate dumb people")
                return False
            print('\n!!! This is not a valid Player ID !!!\n')
            print('Please try again: ')

        # If profile exists            
        else:
            os.system("cls")
            # Print menu
            choice = mainmenu(floor)
            if choice == "play":
                return True
            elif choice == "quit":
                return False
            elif choice == "choose":
                cont_choose_prof = True

# Ivan
def confirmation(floorchosen):
    global floor
    getchoice = True
    choice = pyplus.inputYesNo(f"\nWould you like to proceed to floor {floorchosen}, {profile[1]}? (Yes//No)\n\n>>> ")
    if choice == "yes":
        getchoice = False
        # Set the combat floor
        floor = floorchosen
                                    
    elif choice == "no":
        getchoice = True

    # returns whether need to print menu 
    return getchoice

# Ivan
def mainmenu(floorgiven):
    global done
    global floor
    getchoice = True
    # Print menu with choices
    while getchoice:
        os.system('cls')
        print(f"\nWelcome back {profile[1]}")
        time.sleep(1)
        print("\n+++++++++++++++++++++++++++++++++++++++++++++++++++\n")

        """
        Check the current floor of the player:
        If 0:
            Proceed, Choose profile, Quit
        If <=5 and not 0:
            Proceed, Choose profile, Replay, check, buy upgrades, quit
        If >5:
            Reset, replay, check, buy upgrades, choose profile, quit
        """

        if floorgiven == 0:
            print("\nYou currently have no progress")
            print("Do you want to proceed or change profile?\n")
            choice = pyplus.inputMenu(["Proceed", "Choose Profile", "Quit Game"], "\n",blank=False)

        elif floorgiven <=5 and floorgiven != 0:
            print(f"\nYour current progress is at floor {floor}\nWhat would you like to do?\n")
            choice = pyplus.inputMenu(["Proceed", "Replay", "Check", "Buy Upgrades", "Choose Profile", "Quit Game"], "\n",blank=False)

        else:
            print(f"\nYou have completed the game\nWhat would you like to do?\n")
            choice = pyplus.inputMenu(["Reset", "Replay", "Check", "Buy Upgrades", "Choose Profile", "Quit Game"], "\n",blank=False)

        # Proceed
        if choice == "Proceed":
            os.system("cls")
            # Confirmation
            getchoice = confirmation(floorgiven)
            if not getchoice:
                return "play"
            
        # Replay
        elif choice == "Replay" and floorgiven > 0:
            os.system("cls")
            getfloor = True
            while getfloor:
                choice = pyplus.inputInt("\nEnter floor: ",blank=False,min=0,max=profile[6]-1)
                getfloor = False
                print(f"\nFloor {choice} chosen\nYou will only obtain golds")

                # Confirmation
                getchoice = confirmation(choice)
                if not getchoice:
                    return "play"

        # Check
        elif choice == "Check" and floorgiven > 0:
            os.system("cls")
            checkprofile()

        # Upgrades
        elif choice == "Buy Upgrades" and floorgiven > 0:
            os.system("cls")
            buyupgrades()


        # Choose profile
        elif choice == "Choose Profile":
            os.system("cls")
            return "choose"
    
        # quit game
        elif choice == "Quit Game" and floorgiven >= 0:
            os.system("cls")
            return "quit"

        elif choice == "Reset":
            os.system("cls")
            getchoice = True
            input = pyplus.inputYesNo("\nResetting the game will delete all your progress, however your gold and lvl will remain the same\nAre you sure you want to proceed? (Yes/No) \n>>> ", blank=False)
            if input == "yes":

                # Update the values in profile list and excel sheet
                profile[6] = 0
                update()
                floorgiven = 0
                done = False
                t = threading.Thread(target=animate)
                t.start()
                profilesheet = profileframe.worksheets[0]
                for num, row in enumerate(profilesheet):
                    if num == 0:
                        continue
                    elif int(row[0].value) == profile[0]:
                        row[6].value = profile[6] # progress
                profileframe.save("playerprofile.xlsx")
                done = True
                print("\nDone")
                getchoice = True
                break

            elif input == "no":
                continue

# Ivan                      
def buyupgrades():
    global player_gold
    global critrate
    global done
    upgradesheet = profileframe.worksheets[2]
    profilesheet = profileframe.worksheets[0]
    upgradedict = {}
    upgradelist = []
    # Check if upgrade is empty
    check = True
    if upgrade is None:
        check = False
    else:
        # get list of upgrades from a sentence, comma separated
        upgradelist = [x for x in upgrade.split(",")]

    for num, row in enumerate(upgradesheet):
        if num == 0:
            continue
        else:
            
            # Insert new item in dict with format
            upgradedict[str(row[3].value)] = row[5].value , row[6].value
    
    printupgrades = True

    # Print list of upgrades when true
    while printupgrades:
        os.system("cls")
        readscript("shop")
        time.sleep(1)
        print("\n")
        # Print upgrade list
        print_data_in_table_form("playerprofile.xlsx", 2, "D,F,G")
        time.sleep(1)
        print(f"\nUpgrades aquired: {", ".join(upgradelist)}")
        time.sleep(1)
        print(f"\nAvailable gold: {player_gold}          Current crit rate: {critrate*100}%")
        time.sleep(.5)
        choice = userinputstring("\nPlease choose an upgrade to purchase or type 'back' to exit the shop: ")

        if choice in upgradedict:

            # If player choice is bought and check is true
            if choice in upgradelist:
                print("\nThis item is already in your inventory.\nPlease choose another one")
                time.sleep(1)

            # Choice available or check is false
            elif choice not in upgradelist or not check:

                # get gold required and critrate granted
                gold = upgradedict[choice][0]
                critrate = upgradedict[choice][1]

                input = pyplus.inputYesNo(f"\nAre you sure you want to purchase {choice} with {gold} golds? (Avalaible golds: {player_gold})\n>>> ",blank=False)
                # if player confirms to buy
                if input == "yes" and player_gold>=gold:
                    # Add upgrade to upgradelist
                    upgradelist.append(choice)
                    player_gold -= gold
                    print(f"You have bought {choice}!!")
                    done = False
                    t = threading.Thread(target=animate)
                    t.start()

                    # Update the profile
                    for num, row in enumerate(profilesheet):
                        if num == 0:
                            continue
                        elif row[0].value == ID:
                            row[5].value = player_gold
                            row[7].value += critrate
                            row[8].value = ",".join(upgradelist)

                    profileframe.save("playerprofile.xlsx")
                    update()
                    getprofile(ID, profilesheet)
                    done = True
                    # Update upgrade list
                    upgradelist = [x for x in upgrade.split(",")]
                    

                elif input == "no":
                    continue

                elif input == "yes" and player_gold<gold:
                    print("Not enough gold.")
                    time.sleep(1)


        elif choice.lower() == "back":
            printupgrades = False
            break

        else:
            print("Invalid input!!!")
            time.sleep(1)

# Ivan
def printabilities(use):
    abilitysheet = profileframe.worksheets[1]
    global abilities
    abilities = {}
    print("These are your abilities list:\n")

    # Print out the abilities list
    for num, row in enumerate(abilitysheet):

        # If combat use, only print out abilities for certain floors
        if use == "combat":
            if num == 0:
                continue
            elif int(row[4].value) == floor:
                abilities[row[0].value] = row[2].value, row[3].value 

        # If collection use, print out all unlocked abilities
        elif use == "collection":
            if num == 0:
                continue
            elif int(row[4].value) < floor:
                abilities[row[0].value] = row[2].value, row[3].value 
        
    i = 1
    for x in abilities:
        print(f"{i}. {x} -- {abilities[x][0]} {abilities[x][1]}")
        i += 1
        time.sleep(.2)

# Ivan
def editprofile(id, sheet):
    global profile

    getprofile(id, sheet)
    input = pyplus.inputMenu(["Change name", "Delete profile"], "\n------------------------\n\n\n",blank=False,numbered=True)

    if input == "Change name":
        newname = userinputstring(f"\nOriginal name: {profile[1]} \nEnter the new name: ")
        profile[1] = newname

        # Update name in profile
        for num, row in enumerate(sheet):
            if num == 0:
                continue
            if row[0].value == id:
                row[1].value = newname
        time.sleep(1)
        print("\nProfile saved!!!")
        time.sleep(1)

    elif input == "Delete profile":
        # Delete profile data
        for num, row in enumerate(sheet):
            if row[0].value == ID:
                # Delete row
                sheet.delete_rows(num+1)
                print("\nDeleting profile...")
                time.sleep(1)
                print("Profile deleted!!")
                time.sleep(1)
        return True

# Ivan
def createprofile(sheet):
    global profile
    # Getting basic info from User
    # Make sure no duplicated IDs
    cont_enterid = True
    os.system("cls")
    while cont_enterid:
        id = userinputint("Enter your ID (numbers): ")
        if len(str(id)) != 3:
            print("\nID must be a three-digit number.")
        else:
            if sheet.max_row == 1:
                cont_enterid = False
            else:
                for num, row in enumerate(sheet):
                    if num == 0:
                        continue
                    if row[0].value == id:
                        print("This ID has been taken, please choose another ID.")
                        id = userinputint("Enter your ID: ")
                    else:
                        cont_enterid = False

    username = pyplus.inputStr("Enter your Name: ",blank=False)
    difficultiness = setdifficulty()
    HP = 1000
    lvl = 1
    gold = 300
    upgrade = None
    crit, words, vwords, lose, win, timer, progress = 0, 0, 0, 0, 0, 0, 0
    profile = [id, username, difficultiness, HP, lvl, gold, progress, crit, upgrade, words, vwords, lose, win, timer]
    
    input = pyplus.inputYesNo("\n\nAre you sure you want to create this profile? (Yes//No) \n>>> ",blank=False)
    if input == "yes":
        print("Profile saved!!")
        time.sleep(1)
        print("\n\n")
        return profile
    elif input == "no":
        os.system("cls")
        return False

# Ivan
def setdifficulty():
    global difficulty
    cont = True
    while cont:
        print("\nChoose your difficulty: \nEasy --> 1 \nHard (Case and space sensitive)--> 2")
        choice = userinputint("")
        if choice == 1:
            cont = False
            difficulty = "Easy"
        elif choice == 2:
            cont = False
            difficulty = "Hard"
        else:
            print("Invalid choice!!!\n")
    return difficulty

# Scripts written by Yuven and Xuan Yu
def readscript(type):
    # Start
    if type == "start":
        f = open("Dialogues/Details.txt", "r")
        print(f.read().strip())
        f.close()
        userinputstring("\nPress ENTER key to continue\n")
        os.system("cls")
    
    # Shop
    elif type == "shop":
        f = open("Dialogues/shop.txt", "r")
        print(f.read().strip())
        f.close()
    
    # Game 
    elif type == "game":
        f = open("Dialogues/game.txt", "r")
        print(f.read().strip())
        f.close()

    # Tutorial
    # start tutorial 
    elif type == "0":
        f = open("Dialogues/Tutorial(1).txt", "r")
        readdialogue(f)
        f.close()

    # Tutorial end
    elif type == "0end":
        f = open("Dialogues/Tutorial(2).txt", "r")
        readdialogue(f)
        f.close()
        
    # Quiz
    elif type == "1":
        f = open("Dialogues/Quiz(1).txt", "r")
        readdialogue(f)
        f.close()
    
    # Quiz end
    elif type == "1end":
        f = open("Dialogues/Quiz(2).txt", "r")
        readdialogue(f)
        f.close()

    # Calculator
    elif type == "2":
        f = open("Dialogues/Calculator(1).txt", "r")
        readdialogue(f)
        f.close()
    
    # Calculator end
    elif type == "2end":
        f = open("Dialogues/Calculator(2).txt", "r")
        readdialogue(f)
        f.close()

    # Laptop
    elif type == "3":
        f = open("Dialogues/Laptop(1).txt", "r", encoding="utf-8")
        readdialogue(f)
        f.close()
    
    # Laptop end
    elif type == "3end":
        f = open("Dialogues/Laptop(2).txt", "r")
        readdialogue(f)
        f.close()

    # Calculus
    elif type == "4":
        f = open("Dialogues/Calculus(1).txt", "r")
        readdialogue(f)
        f.close()
    
    # Calculus end
    elif type == "4end":
        f = open("Dialogues/Calculus(2).txt", "r")
        readdialogue(f)
        f.close()

    # Dictionary
    elif type == "5":
        f = open("Dialogues/Dictionary(1).txt", "r")
        readdialogue(f)
        f.close()
    
    # Dictionary end
    elif type == "5end":
        f = open("Dialogues/Dictionary(2).txt", "r")
        readdialogue(f)
        f.close()

    else:
        pass

# Yuven and Xuan Yu
def readdialogue(file):
    for line in file:
            printline = True
            linelist = line.split() # Separate the sentence to a list of strings
            line = line.replace("playername", name)
            for x in linelist: # Check all strings in list

                # If the line contains !Input, get user input and skip the line
                if x == "!Input":
                    printline = False 
                    cont = True  
                    x = linelist[1:] 
                    expectedinput = " ".join(x) 
                    while cont:
                        input = userinputstring("") #get user input 
                        if input == expectedinput: 
                            print("")
                            cont = False
                        elif input == "skip": # pass if user input skip
                            break

                        else:
                            cont = True
                #if the line contains !noinput , continue the scrip
                elif x == "!noinput":  
                    printline = False 
                    userinputstring("") 
                    continue 
                #if the line contains !Game , get user input and continue 
                elif x == "!Game": 
                    printline = False
                    cont = True
                    while cont:
                        input = userinputstring("Type: Challenge accepted\n")
                        if input.lower().strip() == "challenge accepted":
                            cont = False
                            continue
                        else:
                            cont = True

                break
            if printline:
                time.sleep(2)
                print(line)
                time.sleep(1)

# Ivan
def main():
    readscript("start")
    play = True
    play = choooseloadornew()
    
    # Loop play 
    while play:
        os.system("cls")
        print("Type 'skip' to skip the story, else press 'Enter'\n")

        # Enter key to continue, skip to skip story into game
        input = userinputstring("")
        if input.lower() == "skip":
            print("Skipped story....")
        else:
            os.system("cls")
            readscript(str(floor))
        os.system("cls")
        time.sleep(1)
        print("\n\n\n")
        readscript("game")
        print("\n\n\n")
        time.sleep(1)
        play = combat(floor)
    print("\nThanks for playing Keyboard Warrior!!!")

if __name__ == "__main__":
    main()   

#g = Game()

#while g.running:
#    g.welcome.display_title()
#    g.curr_menu.display_menu()
#    g.game_loop()
        
